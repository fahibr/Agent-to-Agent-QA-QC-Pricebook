"""
file.py — Combined testing pipeline for pricebook automation.

Purpose
-------
This script merges two previously separate steps into one runnable flow:

  Step 1 (from main.py)     : Validate HardwareStandard against the master
                              price list (active_price_list.xlsx). Correct
                              mismatched fields (yellow). Products not in
                              master are left unchanged with no highlight.
  Step 2 (from read_file.py): Revise HardwareType (fill blanks with NULL)
                              and sync HardwareStandard ATTYPE1-4 / DHI
                              from HardwareType using locale + category.

How to run
----------
  python file.py          # default: direct Step 1 + Step 2 (no LLM)
  python file.py --agent  # LangChain agent + prompt_v4.txt (uses Azure OpenAI)

Direct mode is the default for testing so Azure rate limits (429) do not
block Excel validation. With --agent, a 429 automatically falls back to
direct mode.
"""

# ======================================================
# Standard library & third-party imports
# ======================================================

import os          # Read env vars and build file paths
import shutil      # Copy the original Excel before writing corrections
import warnings    # Suppress noisy (but harmless) openpyxl warnings

import pandas as pd
from dotenv import load_dotenv
from langchain_openai import AzureChatOpenAI
from langchain.agents import create_agent
from langchain_core.tools import tool
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# openpyxl cannot round-trip some Excel data-validation extensions;
# the warning is noisy but does not affect the cell values we read/write.
warnings.filterwarnings(
    "ignore",
    message="Data Validation extension is not supported and will be removed",
    category=UserWarning,
    module="openpyxl",
)

# ======================================================
# Load Environment (Azure OpenAI credentials)
# ======================================================

# Load key/value pairs from a local .env file into os.environ.
load_dotenv()

# Fail fast if the API key is missing — the agent cannot run without it.
if not os.getenv("AZURE_OPENAI_API_KEY"):
    raise ValueError("AZURE_OPENAI_API_KEY not found in .env file")

# ======================================================
# Shared file / sheet configuration
# ======================================================

# Source pricebook Excel (Step 1 input; also Step 2 fallback input).
PRICEBOOK_EXCEL_PATH = "AD_July2026 (en_HK).xlsx"

# Master / active price list used by Step 1 validation.
MASTER_PRICE_LIST = "active_price_list.xlsx"

# Combined A2A prompt (Agent_1 handoff + Agent_2 QA/QC). Evolved from prompt_v3.
PROMPT_PATH = os.path.join("prompt", "prompt_v4.txt")

# Worksheet names used by both steps.
HARDWARE_STANDARD_SHEET = "HardwareStandard"
HARDWARE_TYPE_SHEET = "HardwareType"

# Locales that may appear in the pricebook filename (Step 2 detection).
LOCALE_LIST = [
    "en_HK",
    "en_SG",
    "en_PH",
    "en_VN",
    "ms_MY",
    "in_ID",
    "zh_TW",
    "zh_CN",
    "th_TH",
]

# Yellow fill = cells corrected / filled from a trusted source.
UPDATED_FILL = PatternFill(
    start_color="FFFF00",
    end_color="FFFF00",
    fill_type="solid",
)

# ======================================================
# Master price list column names (Step 1)
# ======================================================

BrandCode = "Brand Code"
CategoryCode = "Product Category Code"
ProductCategory = "Product Category"
ModelCode = "Model (Item Code)"
Finish = "Finish"
Description = "Description"
Unit = "Unit"
Price = "Project/ Spec Price"

# ======================================================
# Pricebook HardwareStandard column names (Step 1)
# ======================================================

Hdw_Type_Id = "HDW_TYPE_ID"
Product_Line = "PRODUCT_LINE"
Manufacturer = "Manufacturer"
Model = "PRINT_ITEM_DESC"
P_Finish = "PRINT_FINISH"
P_Description = "HDW_TYPE_DESCRIPTION"
P_Unit = "UOM"
P_Price = "PRICE"
P_Attr1 = "ATTR1"

# ======================================================
# HardwareType / HardwareStandard column names (Step 2)
# ======================================================

# HardwareType columns
HT_LOCALE = "Locale"
HT_PRODUCT_CATEGORY = "HDW_TYPE_ID"  # product category code
HT_ATTR1 = "ATTR_TYPE1_ID"
HT_ATTR2 = "ATTR_TYPE2_ID"
HT_ATTR3 = "ATTR_TYPE3_ID"
HT_ATTR4 = "ATTR_TYPE4_ID"
HT_SORT_ORDER = "SORT_ORDER"

# HardwareStandard columns used by Step 2 attribute sync
HWS_PRODUCT_CATEGORY = "HDW_TYPE_ID"  # product category code
HWS_ATTYPE1 = "ATTYPE1"
HWS_ATTYPE2 = "ATTYPE2"
HWS_ATTYPE3 = "ATTYPE3"
HWS_ATTYPE4 = "ATTYPE4"
HWS_DHI = "DHI"  # target for HardwareType.SORT_ORDER

# Mapping: HardwareStandard column <- HardwareType column
ATTR_MAP = [
    (HWS_ATTYPE1, HT_ATTR1),
    (HWS_ATTYPE2, HT_ATTR2),
    (HWS_ATTYPE3, HT_ATTR3),
    (HWS_ATTYPE4, HT_ATTR4),
]

# ======================================================
# Step 1 field-check definitions
# ======================================================

# Each tuple = (friendly name, pricebook column, master column).
# These five fields are compared for every matched Brand+Model+Finish.
FIELD_CHECKS = [
    ("Product Category Code", Hdw_Type_Id, CategoryCode),
    ("Product Category", Product_Line, ProductCategory),
    ("Description", P_Description, Description),
    ("Unit", P_Unit, Unit),
    ("Price", P_Price, Price),
]

# Legacy validation helper columns that must be stripped from output
# so the corrected pricebook stays clean (no extra validation columns).
REMOVED_VALIDATION_COLUMNS = [
    "Validation Status",
    "Mismatched Fields",
    *[f"{field_name} Status" for field_name, _, _ in FIELD_CHECKS],
    *[f"{field_name} (Master)" for field_name, _, _ in FIELD_CHECKS],
]

# ======================================================
# Load Excel data (reloadable for A2A handoff)
# ======================================================

all_sheets = None
HWS_df = None
master_price_list_df = None


def configure_data_sources(
    pricebook_path: str | None = None,
    master_path: str | None = None,
) -> None:
    """
    Load (or reload) pricebook + master DataFrames used by Step 1 tools.

    Called at import when default files exist, and again by run_agent_2 /
    the A2A orchestrator when paths are supplied after Agent_1 handoff.
    """
    global PRICEBOOK_EXCEL_PATH, MASTER_PRICE_LIST
    global all_sheets, HWS_df, master_price_list_df

    next_pricebook = pricebook_path or PRICEBOOK_EXCEL_PATH
    next_master = master_path or MASTER_PRICE_LIST

    loaded_sheets = pd.read_excel(next_pricebook, sheet_name=None)
    loaded_hws = loaded_sheets[HARDWARE_STANDARD_SHEET]
    loaded_master = pd.read_excel(next_master)

    # Commit only after both files load successfully.
    PRICEBOOK_EXCEL_PATH = next_pricebook
    MASTER_PRICE_LIST = next_master
    all_sheets = loaded_sheets
    HWS_df = loaded_hws
    master_price_list_df = loaded_master


def _try_load_default_sources() -> None:
    """Best-effort load for CLI; A2A may configure paths later."""
    try:
        configure_data_sources()
    except FileNotFoundError as exc:
        print(f"[agent_2] Default Excel sources not loaded yet: {exc}")


_try_load_default_sources()


# ======================================================
# Shared path / Excel helpers
# ======================================================

def _corrected_pricebook_path(source_path: str) -> str:
    """
    Build the corrected output filename.

    Example:
      'PL_July2026 (en_HK).xlsx'
        -> 'PL_July2026 (en_HK) - corrected.xlsx'
    """
    base, ext = os.path.splitext(source_path)
    return f"{base} - corrected{ext}"


def _get_header_map(worksheet):
    """
    Map header cell text -> 1-based column index for the first row.

    Used so we can write updates by column name instead of hard-coded
    column letters (which break if columns are reordered).
    """
    return {
        worksheet.cell(1, column_index).value: column_index
        for column_index in range(1, worksheet.max_column + 1)
        if worksheet.cell(1, column_index).value
    }


def _is_missing(value) -> bool:
    """
    Return True when a cell should be treated as empty.

    Covers None, pandas NaN, blank strings, the literal 'nan',
    and the literal 'NULL' (case-insensitive for nan; exact for NULL).
    """
    if value is None or pd.isna(value):
        return True
    text = str(value).strip()
    return text == "" or text.lower() == "nan" or text.upper() == "NULL"


def _normalize_excel_text(value) -> str:
    """
    Strip Excel float artifacts like '31071.0' -> '31071'.

    Excel often stores numeric-looking codes as floats; comparing the
    raw string would falsely report mismatches (e.g. '31071' vs '31071.0').
    """
    text = str(value).strip()
    if text.endswith(".0"):
        try:
            number = float(text)
            # Only strip .0 when the float is a whole number.
            if number == int(number):
                return str(int(number))
        except ValueError:
            # Not a number after all — keep the original text.
            pass
    return text


def _normalize_finish(value) -> str:
    """Normalize finish codes the same way as other Excel text fields."""
    return _normalize_excel_text(value)


def _normalize_price(value):
    """
    Compare prices numerically (rounded) when possible.

    Falls back to stripped text if the value is not numeric
    (e.g. blank, N/A, or a formula remnant).
    """
    try:
        return round(float(str(value).strip()), 6)
    except (TypeError, ValueError):
        return str(value).strip()


def _normalize_value(value):
    """
    Prefer int for whole numbers (e.g. 26.0 -> 26).

    Used by Step 2 when comparing ATTR / SORT_ORDER values so that
    float artifacts do not look like mismatches.
    """
    if _is_missing(value):
        return None
    try:
        number = float(value)
        if number == int(number):
            return int(number)
        return number
    except (TypeError, ValueError):
        return value


# ======================================================
# STEP 1 helpers — master-price-list matching & comparison
# ======================================================

def _make_bm_key(brand, model) -> str:
    """
    Build a Brand|Model lookup key.

    Used as a first-pass filter before finish is considered.
    """
    return (
        f"{_normalize_excel_text(brand)}|"
        f"{_normalize_excel_text(model)}"
    )


def _make_product_key(brand, model, finish) -> str:
    """
    Build a Brand|Model|Finish lookup key.

    This is the unique product identity for Step 1 validation.
    Same item code with a different finish is a different product.
    """
    return (
        f"{_normalize_excel_text(brand)}|"
        f"{_normalize_excel_text(model)}|"
        f"{_normalize_finish(finish)}"
    )


def _find_master_row(row, master_df):
    """
    Return the matching master row for a pricebook row, or None.

    Match order:
      1. Filter master by Brand + Model (BM key).
      2. Prefer exact Brand + Model + Finish (BMF key).
      3. Fall back to normalized finish comparison on BM matches.
      4. If nothing matches -> product does not exist in master (caller skips).
    """
    brand = str(row[Manufacturer]).strip()
    model = str(row[Model]).strip()
    finish = row[P_Finish]

    # First narrow by Brand + Model.
    bm_key = _make_bm_key(brand, model)
    matches = master_df[master_df["BM"] == bm_key]

    if matches.empty:
        # Brand+Model not in master at all.
        return None

    # Prefer the exact Brand+Model+Finish key.
    bmf_key = _make_product_key(brand, model, finish)
    exact_matches = matches[matches["BMF"] == bmf_key]
    if not exact_matches.empty:
        return exact_matches.iloc[0]

    # Fallback: compare normalized finish values one-by-one.
    normalized_finish = _normalize_finish(finish)
    for _, master_row in matches.iterrows():
        if _normalize_finish(master_row[Finish]) == normalized_finish:
            return master_row

    # Brand+Model exists, but this finish does not -> not in master.
    return None


def _compare_field(field_name, pricebook_value, master_value):
    """
    Compare one pricebook field against the master value.

    Returns:
      (is_match: bool, normalized_pricebook, normalized_master)

    Finish and Price use specialized normalizers; all other fields
    are compared as stripped strings.
    """
    if field_name == "Finish":
        pb_value = _normalize_finish(pricebook_value)
        master_value = _normalize_finish(master_value)
        return pb_value == master_value, pb_value, master_value

    if field_name == "Price":
        pb_value = _normalize_price(pricebook_value)
        master_value = _normalize_price(master_value)
        return pb_value == master_value, pb_value, master_value

    pb_value = str(pricebook_value).strip()
    master_value = str(master_value).strip()
    return pb_value == master_value, pb_value, master_value


def _remove_validation_columns(worksheet):
    """
    Delete leftover validation helper columns from HardwareStandard.

    Columns are deleted right-to-left so earlier indices stay valid
    while later ones are removed.
    """
    header_map = _get_header_map(worksheet)
    columns_to_remove = sorted(
        (
            header_map[column_name]
            for column_name in REMOVED_VALIDATION_COLUMNS
            if column_name in header_map
        ),
        reverse=True,  # delete from the right so left indices stay stable
    )

    for column_index in columns_to_remove:
        worksheet.delete_cols(column_index, 1)

    return len(columns_to_remove)


def _format_mismatch_details(mismatch_records):
    """
    Pretty-print per-row mismatch details for the agent / console summary.
    """
    if not mismatch_records:
        return "  (none)"

    lines = []
    for record in mismatch_records:
        lines.append(
            f"  Row {record['row']} | "
            f"Brand: {record['brand']} | "
            f"Model: {record['model']} | "
            f"Finish: {record['finish']}"
        )
        for field in record["fields"]:
            lines.append(f"    - {field['name']} ({field['column']}):")
            lines.append(f"        Pricebook : {field['pricebook']}")
            lines.append(f"        Master    : {field['master']}")
            lines.append("        Status    : Corrected & highlighted yellow")
        lines.append("")
    return "\n".join(lines).rstrip()


def _update_hardware_standard_sheet(path, row_updates):
    """
    Apply Step 1 corrections / yellow highlights to HardwareStandard in-place.

    Parameters
    ----------
    path : str
        Path to the corrected workbook (already copied from the original).
    row_updates : list[dict]
        One dict per HardwareStandard data row (Excel row = index + 2).
        Keys are pricebook column names; values are corrected master values.
        Empty dict = no field corrections for that row.

    Returns
    -------
    dict with sheet / cell update statistics for the summary.
    """
    workbook = load_workbook(path)
    worksheet = workbook[HARDWARE_STANDARD_SHEET]
    original_sheet_names = workbook.sheetnames
    original_max_column = worksheet.max_column

    # Strip any leftover validation helper columns first.
    removed_columns = _remove_validation_columns(worksheet)
    header_map = _get_header_map(worksheet)
    corrected_cells = 0

    # row_updates[0] corresponds to Excel row 2 (row 1 is the header).
    for row_index, corrections in enumerate(row_updates, start=2):
        # Write each mismatched field from the master and paint yellow.
        for pricebook_column, corrected_value in corrections.items():
            if pricebook_column not in header_map:
                continue

            cell = worksheet.cell(
                row_index,
                header_map[pricebook_column],
                value=corrected_value,
            )
            cell.fill = UPDATED_FILL
            corrected_cells += 1

    workbook.save(path)

    return {
        "sheet_names": original_sheet_names,
        "sheet_count": len(original_sheet_names),
        "original_max_column": original_max_column,
        "updated_max_column": worksheet.max_column,
        "removed_validation_columns": removed_columns,
        "rows_processed": len(row_updates),
        "corrected_cells": corrected_cells,
    }


# ======================================================
# STEP 2 helpers — locale / HardwareType lookup / writes
# ======================================================

def _find_locale(filename: str, locales) -> str | None:
    """
    Detect locale code from the pricebook filename.

    Example: 'PL_July2026 (en_HK).xlsx' -> 'en_HK'
    Returns None if no known locale token is present.
    """
    return next((locale for locale in locales if locale in filename), None)


def _build_hardware_type_lookup(ht_df, locale: str) -> dict:
    """
    Build HDW_TYPE_ID -> HardwareType row lookup for one locale.

    Values come from the original HardwareType rows (before NULL fill)
    so ATTR / SORT_ORDER sources are the real typed values, not 'NULL'.
    """
    locale_rows = ht_df[ht_df[HT_LOCALE] == locale]
    lookup = {}
    for _, row in locale_rows.iterrows():
        key = str(row[HT_PRODUCT_CATEGORY]).strip()
        lookup[key] = row
    return lookup


def _apply_updates(worksheet, updates, highlight=True):
    """
    Write a dict of {(excel_row, column_name): value} into a worksheet.

    excel_row is 1-based (row 1 = header).
    Set highlight=False to write values without yellow fill
    (used for HardwareType NULL fills).
    """
    headers = _get_header_map(worksheet)
    count = 0
    for (excel_row, column_name), value in updates.items():
        col = headers.get(column_name)
        if col is None:
            continue
        cell = worksheet.cell(excel_row, col, value=value)
        if highlight:
            cell.fill = UPDATED_FILL
        count += 1
    return count


# ======================================================
# STEP 1 TOOL — Validate against master price list
# ======================================================

@tool
def validate_pricebook():
    """
    STEP 1: Validate HardwareStandard against active_price_list.xlsx.

    Products are uniquely identified by Brand Code + Model + Finish.
    If that combination does not exist in the master price list, the row
    is reported as not in master and left unchanged (no highlight).
    Matched products are compared for exact match on Product Category Code,
    Product Category, Description, Unit, and Price.

    Generates a corrected pricebook file when field mismatches are found.
    Mismatched fields are corrected from the master price list and
    highlighted yellow. The original pricebook file is not modified.
    If all matched products are exact matches, no corrected file is
    generated by this step (Step 2 may still create one).

    Returns a detailed summary including per-row mismatch information
    and a count of products not found in the master.
    """

    # Work on copies so module-level DataFrames stay unchanged.
    hws = HWS_df.fillna("").copy()
    master = master_price_list_df.fillna("").copy()

    # Pre-compute BM and BMF keys on every master row for fast filtering.
    master["BM"] = master.apply(
        lambda row: _make_bm_key(row[BrandCode], row[ModelCode]),
        axis=1,
    )
    master["BMF"] = master.apply(
        lambda row: _make_product_key(
            row[BrandCode],
            row[ModelCode],
            row[Finish],
        ),
        axis=1,
    )

    # One corrections dict per HardwareStandard row (aligned with Excel rows).
    hardware_standard_updates = []

    # Counters / detail lists for the human-readable summary.
    total_checked = 0
    not_in_master = 0
    exact_matches = 0
    partial_mismatches = 0
    corrected_fields = 0
    field_mismatch_counts = {name: 0 for name, _, _ in FIELD_CHECKS}
    mismatch_records = []

    # --------------------------------------------------
    # Walk every HardwareStandard product row
    # --------------------------------------------------
    for row_index, row in hws.iterrows():
        # pandas index 0 == Excel row 2 (row 1 is the header).
        excel_row = row_index + 2
        total_checked += 1

        brand = _normalize_excel_text(row[Manufacturer])
        model = _normalize_excel_text(row[Model])
        finish = _normalize_finish(row[P_Finish])

        master_row = _find_master_row(row, master)

        # --- Not in master: Brand+Model+Finish missing — leave unchanged ---
        if master_row is None:
            not_in_master += 1
            # No field corrections and no highlight.
            hardware_standard_updates.append({})
            continue

        # --- Matched product: compare the five FIELD_CHECKS ---
        mismatched_fields = []
        field_details = []
        corrections = {}

        for field_name, pb_col, master_col in FIELD_CHECKS:
            is_match, pb_value, master_value = _compare_field(
                field_name,
                row[pb_col],
                master_row[master_col],
            )

            if not is_match:
                mismatched_fields.append(field_name)
                field_mismatch_counts[field_name] += 1
                # Queue a write of the master value into the pricebook column.
                corrections[pb_col] = master_row[master_col]
                corrected_fields += 1
                field_details.append(
                    {
                        "name": field_name,
                        "column": pb_col,
                        "pricebook": pb_value,
                        "master": master_value,
                    }
                )

        if mismatched_fields:
            partial_mismatches += 1
            mismatch_records.append(
                {
                    "row": excel_row,
                    "brand": brand,
                    "model": model,
                    "finish": finish,
                    "fields": field_details,
                }
            )
        else:
            exact_matches += 1

        hardware_standard_updates.append(corrections)

    # --------------------------------------------------
    # Write corrected workbook only when fields were corrected
    # --------------------------------------------------
    corrected_path = _corrected_pricebook_path(PRICEBOOK_EXCEL_PATH)
    if corrected_fields > 0:
        # Never mutate the original — copy first, then edit the copy.
        shutil.copy2(PRICEBOOK_EXCEL_PATH, corrected_path)
        update_result = _update_hardware_standard_sheet(
            corrected_path,
            hardware_standard_updates,
        )
        corrected_file_summary = f"""
Corrected pricebook saved:
{corrected_path}
  - All {update_result['sheet_count']} worksheets retained
  - Removed {update_result['removed_validation_columns']} validation columns from HardwareStandard
  - {update_result['corrected_cells']} source cells corrected and highlighted yellow
  - HardwareStandard column count: {update_result['original_max_column']} -> {update_result['updated_max_column']}

Original pricebook unchanged:
{PRICEBOOK_EXCEL_PATH}
"""
    else:
        corrected_file_summary = f"""
Corrected pricebook not generated by Step 1:
  - No field mismatches found among products that exist in the master
  - Products not in master are left unchanged (no highlight)
  - Original pricebook unchanged: {PRICEBOOK_EXCEL_PATH}
  - Step 2 may still create a corrected file for HardwareType / ATTYPE / DHI updates
"""

    field_summary = "\n".join(
        f"{field_name:24}: {count}"
        for field_name, count in field_mismatch_counts.items()
    )
    mismatch_details = _format_mismatch_details(mismatch_records)

    summary = f"""
========== STEP 1: Validation Completed ==========

Total Products Checked : {total_checked}
Exact Matches          : {exact_matches}
Partial Mismatches     : {partial_mismatches}
Not In Master          : {not_in_master}
Fields Corrected       : {corrected_fields} (highlighted yellow in pricebook)

Field Mismatch Counts:
{field_summary}

Mismatch Details (corrected in pricebook):
{mismatch_details}
{corrected_file_summary}
"""

    return summary


# ======================================================
# STEP 2 TOOL — HardwareType NULL fill + ATTYPE / DHI sync
# ======================================================

@tool
def revise_pricebook_attributes():
    """
    STEP 2: Revise HardwareType and HardwareStandard attribute columns.

    Uses the Step 1 corrected pricebook when it already exists; otherwise
    uses the original pricebook. Then:

      1. Detect locale from the filename (e.g. en_HK).
      2. HardwareType: fill missing cells with \"NULL\" (no highlight).
      3. HardwareStandard: match HardwareType by locale + product category
         code (HardwareType.HDW_TYPE_ID / HardwareStandard.HDW_TYPE_ID), then:
           - fill missing ATTYPE1-4 from ATTR_TYPE*_ID
           - correct ATTYPE1-4 when values differ from ATTR_TYPE*_ID
           - set ATTYPE to \"NULL\" when HardwareType ATTR is blank
           - copy / correct DHI from SORT_ORDER
         Highlight filled/corrected ATTR/DHI cells yellow.
      4. HardwareStandard: fill any remaining blank cells with \"NULL\"
         (no highlight).

    Retains all worksheets. Does not modify the original pricebook file.
    Returns a detailed summary of fills, corrections, and the output path.
    """

    # Prefer Step 1 output when present; otherwise start from the original.
    corrected_path = _corrected_pricebook_path(PRICEBOOK_EXCEL_PATH)
    if os.path.exists(corrected_path):
        source_path = corrected_path
        source_note = "Step 1 corrected file"
    else:
        source_path = PRICEBOOK_EXCEL_PATH
        source_note = "original pricebook (Step 1 produced no corrected file)"

    return _revise_pricebook_core(source_path, source_note=source_note)


def _revise_pricebook_core(source_path: str, locales=None, source_note: str = ""):
    """
    Core Step 2 logic (also callable directly for non-agent testing).

    Parameters
    ----------
    source_path : str
        Excel file to read (original or Step 1 corrected).
    locales : list[str] | None
        Locale tokens to search for in the filename.
    source_note : str
        Short label describing which input was chosen (for the summary).

    Returns
    -------
    str
        Human-readable summary of the revision.
    """
    locales = locales or LOCALE_LIST

    # Load all sheets; we only mutate HardwareType + HardwareStandard.
    all_sheets_local = pd.read_excel(source_path, sheet_name=None)
    ht_df = all_sheets_local[HARDWARE_TYPE_SHEET].copy()
    hws_df = all_sheets_local[HARDWARE_STANDARD_SHEET].copy()

    # Locale must appear in the filename (e.g. en_HK).
    locale = _find_locale(source_path, locales)
    if not locale:
        raise ValueError(f"No matching locale found in filename: {source_path}")

    # Snapshot HardwareType BEFORE writing NULL so lookup values stay real.
    ht_lookup_source = ht_df.copy()
    type_lookup = _build_hardware_type_lookup(ht_lookup_source, locale)
    if not type_lookup:
        raise ValueError(f"No HardwareType rows found for locale {locale}")

    # Queues of openpyxl cell writes: {(excel_row, column_name): value}
    ht_updates = {}
    hws_updates = {}

    # --------------------------------------------------
    # HardwareType: fill missing values with the literal "NULL"
    # --------------------------------------------------
    for index, row in ht_df.iterrows():
        excel_row = index + 2  # pandas 0 -> Excel row 2
        for column_name in ht_df.columns:
            if _is_missing(row[column_name]):
                ht_updates[(excel_row, column_name)] = "NULL"

    # Keep the in-memory frame consistent (fillna upcasts numeric cols).
    ht_df = ht_df.fillna("NULL")

    # --------------------------------------------------
    # HardwareStandard: map by locale + product category
    # --------------------------------------------------
    filled_counts = {col: 0 for col, _ in ATTR_MAP}
    corrected_counts = {col: 0 for col, _ in ATTR_MAP}
    filled_counts[HWS_DHI] = 0
    corrected_counts[HWS_DHI] = 0
    rows_updated = 0
    mapped_types = set()
    unmapped_types = set()
    mismatch_examples = []

    for index, row in hws_df.iterrows():
        excel_row = index + 2
        product_category = str(row[HWS_PRODUCT_CATEGORY]).strip()
        ht_row = type_lookup.get(product_category)

        # Product category has no HardwareType row for this locale.
        if ht_row is None:
            unmapped_types.add(product_category)
            continue

        mapped_types.add(product_category)
        row_changed = False

        # ATTYPE1-4: fill if missing, correct if wrong, or clear to NULL
        # when HardwareType ATTR_TYPE*_ID is blank (e.g. FN ATTR_TYPE3_ID).
        for hws_col, ht_col in ATTR_MAP:
            expected = _normalize_value(ht_row[ht_col])
            current = _normalize_value(row[hws_col])

            # HardwareType ATTR is blank -> HardwareStandard ATTYPE must be NULL.
            if expected is None:
                if current is None:
                    continue
                write_value = "NULL"
                # ATTYPE cols are often float64; cast so "NULL" can be stored.
                if hws_df[hws_col].dtype != object:
                    hws_df[hws_col] = hws_df[hws_col].astype(object)
                hws_df.at[index, hws_col] = write_value
                hws_updates[(excel_row, hws_col)] = write_value
                corrected_counts[hws_col] += 1
                if len(mismatch_examples) < 20:
                    mismatch_examples.append(
                        f"  Row {excel_row} | {product_category} | {hws_col}: "
                        f"{current} -> NULL"
                    )
                row_changed = True
                continue

            if current == expected:
                continue

            # Update in-memory frame and queue an Excel write.
            hws_df.at[index, hws_col] = expected
            hws_updates[(excel_row, hws_col)] = expected
            if current is None:
                filled_counts[hws_col] += 1
            else:
                corrected_counts[hws_col] += 1
                # Keep a small sample of mismatches for the summary.
                if len(mismatch_examples) < 20:
                    mismatch_examples.append(
                        f"  Row {excel_row} | {product_category} | {hws_col}: "
                        f"{current} -> {expected}"
                    )
            row_changed = True

        # DHI: fill if missing, or correct if value differs from SORT_ORDER.
        expected_dhi = _normalize_value(ht_row[HT_SORT_ORDER])
        if expected_dhi is not None:
            current_dhi = (
                _normalize_value(row[HWS_DHI]) if HWS_DHI in hws_df.columns else None
            )
            if current_dhi != expected_dhi:
                hws_df.at[index, HWS_DHI] = expected_dhi
                hws_updates[(excel_row, HWS_DHI)] = expected_dhi
                if current_dhi is None:
                    filled_counts[HWS_DHI] += 1
                else:
                    corrected_counts[HWS_DHI] += 1
                    if len(mismatch_examples) < 20:
                        mismatch_examples.append(
                            f"  Row {excel_row} | {product_category} | {HWS_DHI}: "
                            f"{current_dhi} -> {expected_dhi}"
                        )
                row_changed = True

        if row_changed:
            rows_updated += 1

    # --------------------------------------------------
    # HardwareStandard: fill any remaining blank cells with "NULL"
    # (same rule as HardwareType; no yellow highlight)
    # --------------------------------------------------
    hws_null_updates = {}
    for index, row in hws_df.iterrows():
        excel_row = index + 2
        for column_name in hws_df.columns:
            key = (excel_row, column_name)
            # Prefer ATTR/DHI sync values already queued for this cell.
            if key in hws_updates:
                continue
            if _is_missing(row[column_name]):
                hws_null_updates[key] = "NULL"

    # --------------------------------------------------
    # Write both sheets into the corrected pricebook path
    # --------------------------------------------------
    output_path = _corrected_pricebook_path(PRICEBOOK_EXCEL_PATH)

    # If we are not already editing the corrected file, copy the source first.
    # (When source_path IS the corrected file, we edit it in place.)
    if os.path.abspath(source_path) != os.path.abspath(output_path):
        shutil.copy2(source_path, output_path)

    workbook = load_workbook(output_path)
    # HardwareType NULL fills: write values but do NOT highlight.
    ht_cells = _apply_updates(
        workbook[HARDWARE_TYPE_SHEET],
        ht_updates,
        highlight=False,
    )
    # HardwareStandard ATTYPE / DHI updates: write + yellow highlight.
    hws_attr_cells = _apply_updates(
        workbook[HARDWARE_STANDARD_SHEET],
        hws_updates,
        highlight=True,
    )
    # HardwareStandard remaining blanks: write "NULL" without highlight.
    hws_null_cells = _apply_updates(
        workbook[HARDWARE_STANDARD_SHEET],
        hws_null_updates,
        highlight=False,
    )
    workbook.save(output_path)

    mismatch_block = (
        "\n".join(mismatch_examples)
        if mismatch_examples
        else "  (none)"
    )
    if sum(corrected_counts.values()) > len(mismatch_examples):
        mismatch_block += (
            f"\n  ... and {sum(corrected_counts.values()) - len(mismatch_examples)} more"
        )

    summary = f"""
========== STEP 2: Locale HardwareType / HardwareStandard revision ==========

Locale                    : {locale}
Source                    : {source_path}
Source note               : {source_note or 'n/a'}
Output                    : {output_path}
Sheets retained           : {len(workbook.sheetnames)}

HardwareType
  Rows                    : {len(ht_df)}
  Cells filled with NULL  : {ht_cells}

HardwareStandard
  Rows                    : {len(hws_df)}
  Rows updated (ATTR/DHI) : {rows_updated}
  Mapped categories       : {', '.join(sorted(mapped_types)) or '(none)'}
  Unmapped categories     : {', '.join(sorted(unmapped_types)) or '(none)'}

  Filled (was missing):
    ATTYPE1 : {filled_counts[HWS_ATTYPE1]}
    ATTYPE2 : {filled_counts[HWS_ATTYPE2]}
    ATTYPE3 : {filled_counts[HWS_ATTYPE3]}
    ATTYPE4 : {filled_counts[HWS_ATTYPE4]}
    DHI     : {filled_counts[HWS_DHI]}

  Corrected (had wrong value):
    ATTYPE1 : {corrected_counts[HWS_ATTYPE1]}
    ATTYPE2 : {corrected_counts[HWS_ATTYPE2]}
    ATTYPE3 : {corrected_counts[HWS_ATTYPE3]}
    ATTYPE4 : {corrected_counts[HWS_ATTYPE4]}
    DHI     : {corrected_counts[HWS_DHI]}

  ATTR/DHI cells updated  : {hws_attr_cells} (highlighted yellow)
  Blank cells filled NULL : {hws_null_cells} (no highlight)

Mismatch corrections (sample):
{mismatch_block}

Mapping rule:
  HardwareType[{HT_LOCALE}={locale} + {HT_PRODUCT_CATEGORY}]
    -> HardwareStandard[{HWS_PRODUCT_CATEGORY}]
  {HT_SORT_ORDER} -> {HWS_DHI}
  {HT_ATTR1}/{HT_ATTR2}/{HT_ATTR3}/{HT_ATTR4} -> {HWS_ATTYPE1}/{HWS_ATTYPE2}/{HWS_ATTYPE3}/{HWS_ATTYPE4}
""".strip()

    return summary


# ======================================================
# Agent Setup (Azure OpenAI + both Step tools)
# ======================================================

# max_retries: OpenAI SDK backs off and retries on 429 rate limits.
# Raise this if swedencentral gpt-4o-mini still hits TPM/RPM caps.
model = AzureChatOpenAI(
    azure_deployment=os.getenv("DEPLOYMENT_NAME_2"),
    api_key=os.getenv("AZURE_OPENAI_API_KEY"),
    api_version=os.getenv("API_VERSION", "2025-01-01-preview"),
    azure_endpoint=os.getenv(
        "AZURE_OPENAI_ENDPOINT",
        "https://aa-openai-sweden.openai.azure.com/",
    ),
    max_retries=int(os.getenv("AZURE_OPENAI_MAX_RETRIES", "8")),
    timeout=float(os.getenv("AZURE_OPENAI_TIMEOUT", "120")),
)

# Agent can call Agent_2 Step 1 then Step 2 (prompt_v4 documents full A2A order).
tools = [
    validate_pricebook,            # Step 1
    revise_pricebook_attributes,   # Step 2
]

agent = create_agent(
    model=model,
    tools=tools,
)


# ======================================================
# Prompt loader + main entrypoint
# ======================================================

def load_prompt(path: str = PROMPT_PATH) -> str:
    """Read the agent instruction prompt from disk (UTF-8)."""
    with open(path, encoding="utf-8") as prompt_file:
        return prompt_file.read().strip()


def run_pipeline_direct():
    """
    Run Step 1 then Step 2 without calling Azure OpenAI.

    Use this when testing the Excel logic, or when the deployment
    is rate-limited (429). The tools do the real work; the LLM only
    orchestrates and summarizes.
    """
    print("Running in direct mode (no LLM) — Step 1 then Step 2...\n")
    step1_summary = validate_pricebook.invoke({})
    print(step1_summary)
    print("\n" + "=" * 60 + "\n")
    step2_summary = revise_pricebook_attributes.invoke({})
    print(step2_summary)
    return step1_summary, step2_summary


def run_pipeline_with_agent():
    """
    Run the LangChain agent using prompt_v4.txt (uses Azure OpenAI).

    On rate-limit (429), falls back to direct Step 1 + Step 2 so the
    Excel pipeline still completes.

    Agent_2 only has validate_pricebook + revise_pricebook_attributes, so
    the full A2A prompt is scoped to the Agent_2 section at runtime.
    """
    from openai import RateLimitError

    full_prompt = load_prompt()
    agent2_prompt = f"""
You are Agent_2 in the A2A pricebook QA/QC pipeline.
Agent_1 has already completed HANDOFF (active_price_list.xlsx is ready).
Do NOT call Agent_1 tools (load_price_list, remove_discontinued_models,
export_active_list, list_updated_products, list_new_products).

Call ONLY these tools in order:
1) validate_pricebook
2) revise_pricebook_attributes

Follow the AGENT_2 section of the pipeline prompt below.
Final reply: one short Step 1 summary, then one short Step 2 summary.
Do not duplicate the same text twice.

================================================================================
PIPELINE PROMPT (prompt_v4)
================================================================================
{full_prompt}
""".strip()

    try:
        response = agent.invoke(
            {
                "messages": [
                    (
                        "human",
                        agent2_prompt,
                    )
                ]
            }
        )
        content = response["messages"][-1].content
        print(content)
        # Single combined LLM reply — avoid duplicating into step1 and step2.
        return content, ""
    except RateLimitError as exc:
        print(
            "Azure OpenAI rate limit (429). Falling back to direct mode "
            "(no LLM) so Step 1 + Step 2 still run.\n"
            f"Details: {exc}\n"
        )
        return run_pipeline_direct()


def run_agent_2(
    pricebook_path: str | None = None,
    master_path: str | None = None,
    use_llm: bool = False,
) -> dict:
    """
    Run Agent 2 (pricebook QA/QC + attribute revise) and return a QA_RESULT payload.

    Reloads Excel sources from the given paths (typically Agent_1's
    active_price_list.xlsx handoff artifact as master_path).
    """
    configure_data_sources(
        pricebook_path=pricebook_path,
        master_path=master_path,
    )

    if HWS_df is None or master_price_list_df is None:
        return {
            "status": "failed",
            "error": "Pricebook or master data failed to load.",
            "pricebook_path": PRICEBOOK_EXCEL_PATH,
            "master_path": MASTER_PRICE_LIST,
            "step1_summary": "",
            "step2_summary": "",
            "corrected_path": None,
        }

    try:
        if use_llm:
            step1_summary, step2_summary = run_pipeline_with_agent()
        else:
            step1_summary, step2_summary = run_pipeline_direct()
    except Exception as exc:  # noqa: BLE001
        return {
            "status": "failed",
            "error": str(exc),
            "pricebook_path": os.path.abspath(PRICEBOOK_EXCEL_PATH),
            "master_path": os.path.abspath(MASTER_PRICE_LIST),
            "step1_summary": "",
            "step2_summary": "",
            "corrected_path": None,
        }

    corrected_path = _corrected_pricebook_path(PRICEBOOK_EXCEL_PATH)
    if not os.path.isfile(corrected_path):
        corrected_path = None
    else:
        corrected_path = os.path.abspath(corrected_path)

    return {
        "status": "completed",
        "pricebook_path": os.path.abspath(PRICEBOOK_EXCEL_PATH),
        "master_path": os.path.abspath(MASTER_PRICE_LIST),
        "step1_summary": step1_summary if isinstance(step1_summary, str) else str(step1_summary),
        "step2_summary": step2_summary if isinstance(step2_summary, str) else str(step2_summary),
        "corrected_path": corrected_path,
    }


if __name__ == "__main__":
    import sys

    # Ensure default workbooks are available for CLI runs.
    if HWS_df is None or master_price_list_df is None:
        configure_data_sources()

    # file.py is for testing the combined pipeline.
    # Default = direct (no LLM) to avoid Azure 429 rate limits.
    # Use the agent only when explicitly requested:
    #   python file.py --agent
    use_agent = "--agent" in sys.argv

    if use_agent:
        run_pipeline_with_agent()
    else:
        run_pipeline_direct()

